from __future__ import annotations

import json
import re
import unicodedata
from collections.abc import Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sf_report_agent.reports.transforms import clean_html_cells


def _slug(value: str) -> str:
    folded = "".join(
        char
        for char in unicodedata.normalize("NFKD", value.lower())
        if not unicodedata.combining(char)
    )
    return re.sub(r"[^a-z0-9]+", "-", folded).strip("-")[:60] or "reporte"


def _metadata_rows(metadata: dict[str, Any]) -> pd.DataFrame:
    rows = []
    for key, value in metadata.items():
        rendered = (
            json.dumps(value, ensure_ascii=False, default=str)
            if isinstance(value, (dict, list))
            else value
        )
        rows.append({"campo": key, "valor": rendered})
    return pd.DataFrame(rows)


def _style_worksheet(worksheet: Any, *, metadata_sheet: bool = False) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 30

    for index, column in enumerate(worksheet.iter_cols(), start=1):
        values = [str(cell.value) for cell in column if cell.value is not None]
        longest = max((len(value) for value in values), default=0)
        if metadata_sheet and index == 2:
            width = 80
            for cell in column[1:]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        else:
            width = min(max(longest + 2, 12), 34)
        worksheet.column_dimensions[get_column_letter(index)].width = width

    headers = {str(cell.value): cell.column for cell in worksheet[1] if cell.value}
    for header, number_format in {"Edad": "0", "Importe": "#,##0.00"}.items():
        column_index = headers.get(header)
        if column_index is None:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column_index).number_format = number_format


def export_report(
    dataframe: pd.DataFrame,
    *,
    task_id: int,
    title: str,
    artifacts_dir: Path,
    output_formats: Sequence[str],
    metadata: dict[str, Any],
    warnings: list[str],
    generated_at: datetime | None = None,
) -> list[Path]:
    dataframe = clean_html_cells(dataframe)
    generated_at = generated_at or datetime.now(UTC)
    directory = artifacts_dir / "reports"
    directory.mkdir(parents=True, exist_ok=True)
    stamp = generated_at.strftime("%Y%m%dT%H%M%SZ")
    stem = f"task_{task_id}_{_slug(title)}_{stamp}"
    paths: list[Path] = []

    if "csv" in output_formats:
        csv_path = directory / f"{stem}.csv"
        dataframe.to_csv(csv_path, index=False)
        paths.append(csv_path)
    if "xlsx" in output_formats:
        xlsx_path = directory / f"{stem}.xlsx"
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            dataframe.to_excel(writer, sheet_name="datos", index=False)
            _metadata_rows(metadata).to_excel(writer, sheet_name="metadata", index=False)
            if warnings:
                pd.DataFrame({"warning": warnings}).to_excel(
                    writer, sheet_name="warnings", index=False
                )
            _style_worksheet(writer.sheets["datos"])
            _style_worksheet(writer.sheets["metadata"], metadata_sheet=True)
            if warnings:
                _style_worksheet(writer.sheets["warnings"], metadata_sheet=True)
        paths.append(xlsx_path)
    return paths


def write_run_metadata(
    *,
    task_id: int,
    artifacts_dir: Path,
    metadata: dict[str, Any],
    variant_id: str | None = None,
    generated_at: datetime | None = None,
) -> Path:
    generated_at = generated_at or datetime.now(UTC)
    directory = artifacts_dir / "runs"
    directory.mkdir(parents=True, exist_ok=True)
    stamp = generated_at.strftime("%Y%m%dT%H%M%SZ")
    variant_suffix = f"_{_slug(variant_id)}" if variant_id else ""
    path = directory / f"task_{task_id}{variant_suffix}_{stamp}.json"
    path.write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    return path
