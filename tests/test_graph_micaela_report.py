import csv
import json
import sqlite3
from pathlib import Path
from typing import Any

import openpyxl

from conftest import FakeSalesforceClient, create_source_database, write_field_mapping
from sf_report_agent.config import Settings
from sf_report_agent.db.run_repository import ReportRunRepository
from sf_report_agent.db.task_reader import TaskReader
from sf_report_agent.graph.app import ReportAgentRunner
from sf_report_agent.graph.nodes import AgentServices
from sf_report_agent.models.execution_result import ExecutionResult
from sf_report_agent.models.task import ExternalTask


class RecurringDonationSalesforceClient(FakeSalesforceClient):
    def describe_object(self, object_name: str) -> dict[str, Any]:
        if object_name == "Campaign":
            return {
                "fields": [
                    {"name": "Id", "label": "ID de campaña", "type": "id"},
                    {"name": "Name", "label": "Nombre", "type": "string"},
                ]
            }
        if object_name == "Contact":
            return {
                "fields": [
                    {"name": "Id", "label": "ID del contacto", "type": "id"},
                    {"name": "Name", "label": "Nombre", "type": "string"},
                    {"name": "FirstName", "label": "Nombre", "type": "string"},
                    {"name": "LastName", "label": "Apellido", "type": "string"},
                    {
                        "name": "Birthdate",
                        "label": "Fecha de nacimiento",
                        "type": "date",
                    },
                    {"name": "MailingCity", "label": "Ciudad", "type": "string"},
                    {"name": "MailingState", "label": "Provincia", "type": "string"},
                    {
                        "name": "OtherState",
                        "label": "Otra provincia",
                        "type": "string",
                    },
                    {"name": "MailingCountry", "label": "País", "type": "string"},
                ]
            }
        if object_name != "npe03__Recurring_Donation__c":
            return super().describe_object(object_name)
        fields = [
            {"name": "Id", "label": "ID de donación recurrente", "type": "id"},
            {"name": "Name", "label": "Nombre", "type": "string"},
            {"name": "npe03__Amount__c", "label": "Importe", "type": "currency"},
            {
                "name": "npsp__Status__c",
                "label": "Estado",
                "type": "picklist",
                "picklistValues": [
                    {"value": "Active", "label": "Activo", "active": True},
                    {"value": "Closed", "label": "Cerrado", "active": True},
                ],
            },
            {"name": "npsp__StartDate__c", "label": "Fecha inicial", "type": "date"},
            {
                "name": "npsp__EndDate__c",
                "label": "Fecha de finalización",
                "type": "date",
            },
            {
                "name": "npe03__Date_Established__c",
                "label": "Fecha establecida",
                "type": "date",
            },
            {"name": "Fecha_de_alta__c", "label": "Fecha de alta", "type": "date"},
            {
                "name": "npe03__Contact__c",
                "label": "Contacto",
                "type": "reference",
                "referenceTo": ["Contact"],
                "relationshipName": "npe03__Contact__r",
            },
            {
                "name": "Campa_a_de_origen__c",
                "label": "Campaña de origen",
                "type": "reference",
                "referenceTo": ["Campaign"],
                "relationshipName": "Campa_a_de_origen__r",
            },
            {
                "name": "Campa_a_Principal__c",
                "label": "Campaña Principal de Origen",
                "type": "string",
                "filterable": True,
            },
            {
                "name": "npe03__Recurring_Donation_Campaign__c",
                "label": "Campaña para las donaciones futuras",
                "type": "reference",
                "referenceTo": ["Campaign"],
                "relationshipName": "npe03__Recurring_Donation_Campaign__r",
            },
        ]
        return {
            "fields": [
                {
                    "label": field["name"],
                    "referenceTo": [],
                    "relationshipName": None,
                    **field,
                }
                for field in fields
            ]
        }

    def query_all(self, soql: str) -> list[dict[str, Any]]:
        self.queried_soql.append(soql)
        return [
            {
                "attributes": {"type": "npe03__Recurring_Donation__c"},
                "Id": "a0R000000000001",
                "npe03__Amount__c": 1500,
                "npsp__Status__c": "Active",
                "npsp__StartDate__c": "2026-02-15",
                "npsp__EndDate__c": None,
                "npe03__Date_Established__c": "2026-02-10",
                "Fecha_de_alta__c": "2026-02-15",
                "npe03__Contact__r": {
                    "Name": "Persona Uno",
                    "FirstName": "Persona",
                    "LastName": "Uno",
                    "Birthdate": "1990-01-01",
                    "MailingCity": "Córdoba",
                    "MailingState": "Córdoba",
                    "OtherState": None,
                    "MailingCountry": "Argentina",
                },
                "Campa_a_Principal__c": (
                    '<a href="/campaign/redes">[IND] Redes Sociales</a>'
                    if "[IND] Redes Sociales" in soql
                    else '<a href="/campaign/pauta">[IND] Campañas Pauta Digital</a>'
                ),
                "Campa_a_de_origen__c": self.campaign_ids[0],
                "Campa_a_de_origen__r": {"Name": "[IND] Campañas Pauta Digital"},
                "npe03__Recurring_Donation_Campaign__c": self.campaign_ids[1],
                "npe03__Recurring_Donation_Campaign__r": {
                    "Name": "[IND] Redes Sociales"
                },
            }
        ]


class NonFilterableMainCampaignSalesforceClient(RecurringDonationSalesforceClient):
    def describe_object(self, object_name: str) -> dict[str, Any]:
        description = super().describe_object(object_name)
        if object_name == "npe03__Recurring_Donation__c":
            field = next(
                value
                for value in description["fields"]
                if value["name"] == "Campa_a_Principal__c"
            )
            field["filterable"] = False
        return description


class ReportCreationSalesforceClient(FakeSalesforceClient):
    def __init__(self) -> None:
        super().__init__()
        self.create_attempts = 0

    def create_report(self, **kwargs: Any) -> None:
        self.create_attempts += 1
        raise RuntimeError("permiso insuficiente")


def _settings(
    tmp_path: Path,
    source_db: Path,
    *,
    mapping_path: Path | None = None,
    business_semantics_path: Path | None = None,
    allow_salesforce_report_create: bool = False,
) -> Settings:
    return Settings(
        source_db_path=source_db,
        worker_db_path=tmp_path / "worker.db",
        artifacts_dir=tmp_path / "artifacts",
        field_mapping_path=mapping_path,
        model_provider="ollama",
        ollama_model="gemma4:e2b-mlx",
        ollama_base_url="http://127.0.0.1:11434",
        ollama_temperature=0,
        salesforce_username="user@example.org",
        salesforce_password="secret",
        salesforce_security_token="token",
        salesforce_domain="login",
        sf_read_only=True,
        max_export_rows=50_000,
        require_human_approval_for_pii=True,
        log_pii=False,
        update_source_task=False,
        allow_report_without_person_fields=False,
        allow_salesforce_report_create=allow_salesforce_report_create,
        business_semantics_path=business_semantics_path,
    )


def _run(
    tmp_path: Path,
    task: ExternalTask,
    fake_salesforce: FakeSalesforceClient,
    *,
    mapping_path: Path,
    business_semantics_path: Path | None = None,
    dry_run: bool = False,
    allow_salesforce_report_create: bool = False,
) -> tuple[ExecutionResult, Settings]:
    tmp_path.mkdir(parents=True, exist_ok=True)
    source_db = tmp_path / "source.db"
    create_source_database(source_db, task)
    settings = _settings(
        tmp_path,
        source_db,
        mapping_path=mapping_path,
        business_semantics_path=business_semantics_path,
        allow_salesforce_report_create=allow_salesforce_report_create,
    )
    services = AgentServices(
        settings=settings,
        task_reader=TaskReader(source_db),
        run_repository=ReportRunRepository(settings.worker_db_path),
        salesforce_client=fake_salesforce,
        ollama_client=None,
    )
    return ReportAgentRunner(services).run(task.id, dry_run=dry_run), settings


def test_full_graph_with_mock_salesforce_generates_artifacts(
    tmp_path: Path,
    micaela_task: ExternalTask,
    fake_salesforce: FakeSalesforceClient,
) -> None:
    result, _ = _run(
        tmp_path,
        micaela_task,
        fake_salesforce,
        mapping_path=write_field_mapping(tmp_path / "mapping.json"),
    )

    assert result.status == "done_pending_approval"
    assert result.row_count == 2
    assert result.soql.startswith("SELECT ")
    assert "Contact.Name" in result.soql
    assert "LeadSource IN ('amplify', 'orgánico web')" in result.soql
    assert " OR " in result.soql
    assert fake_salesforce.queried_soql
    csv_path = next(Path(path) for path in result.artifacts if path.endswith(".csv"))
    xlsx_path = next(Path(path) for path in result.artifacts if path.endswith(".xlsx"))
    metadata_path = next(
        Path(path) for path in result.artifacts if "/runs/" in path and path.endswith(".json")
    )
    assert csv_path.exists()
    assert xlsx_path.exists()
    assert metadata_path.exists()
    assert openpyxl.load_workbook(xlsx_path).sheetnames == ["datos", "metadata", "warnings"]
    assert "requiere aprobación humana" in result.response_text
    assert "informe de altas 2026" in result.response_text
    assert "[IND] Campañas Pauta Digital" in result.response_text
    assert "amplify, orgánico web" in result.response_text


def test_task_23_uses_recurring_donation_mapping_and_real_fields(
    tmp_path: Path,
    micaela_task: ExternalTask,
) -> None:
    task_23 = micaela_task.model_copy(update={"id": 23})
    mapping_path = Path(__file__).parents[1] / "config" / "field_mapping.json"
    semantics_path = Path(__file__).parents[1] / "config" / "business_semantics.yaml"
    salesforce = RecurringDonationSalesforceClient()

    result, settings = _run(
        tmp_path,
        task_23,
        salesforce,
        mapping_path=mapping_path,
        business_semantics_path=semantics_path,
    )

    assert settings.field_mapping_path == mapping_path
    assert result.status == "done_pending_approval"
    assert len(result.variants) == 2
    by_id = {variant.variant_id: variant for variant in result.variants}
    pauta = by_id["ind_campanas_pauta_digital"]
    redes = by_id["ind_redes_sociales"]
    assert "Campa_a_Principal__c LIKE '%[IND] Campañas Pauta Digital%'" in pauta.soql
    assert "Campa_a_Principal__c LIKE '%[IND] Redes Sociales%'" in redes.soql
    for variant in result.variants:
        assert "FROM npe03__Recurring_Donation__c" in variant.soql
        assert "npe03__Date_Established__c >= 2026-01-01" in variant.soql
        assert "npe03__Date_Established__c < 2027-01-01" in variant.soql
        assert "npe03__Contact__c != NULL" in variant.soql
        assert "CampaignId" not in variant.soql
        assert "CloseDate" not in variant.soql
        where = variant.soql.split("WHERE", 1)[1]
        assert "npe03__Recurring_Donation_Campaign__c" not in where
        assert "Campa_a_de_recupero__c" not in where
    for relationship_field in (
        "npe03__Contact__r.FirstName",
        "npe03__Contact__r.LastName",
        "npe03__Contact__r.Birthdate",
        "npe03__Contact__r.MailingState",
        "npe03__Contact__r.OtherState",
        "Campa_a_de_origen__r.Name",
    ):
        assert all(relationship_field in variant.soql for variant in result.variants)
    assert salesforce.queried_soql == [variant.soql for variant in result.variants]
    assert "informe separado por cada segmento de negocio" in result.response_text
    assert "[IND] Campañas Pauta Digital" in result.response_text
    assert "[IND] Redes Sociales" in result.response_text
    assert "Campaña para las donaciones futuras" not in result.response_text
    assert "Campañas combinadas" not in result.response_text
    assert all(len(variant.artifacts) == 3 for variant in result.variants)

    pauta_csv = next(Path(path) for path in pauta.artifacts if path.endswith(".csv"))
    pauta_xlsx = next(Path(path) for path in pauta.artifacts if path.endswith(".xlsx"))
    expected_headers = [
        "Contacto: Nombre",
        "Contacto: Apellido",
        "Contacto: Fecha de nacimiento",
        "Edad",
        "Provincia",
        "Fecha establecida",
        "Estado",
        "Importe",
        "Fecha de finalización",
        "Campaña Principal de Origen",
        "Campaña de origen",
    ]
    with pauta_csv.open(encoding="utf-8", newline="") as handle:
        csv_rows = list(csv.DictReader(handle))
    assert list(csv_rows[0]) == expected_headers
    assert csv_rows[0]["Estado"] == "Activo"
    assert csv_rows[0]["Campaña Principal de Origen"] == (
        "[IND] Campañas Pauta Digital"
    )
    assert "<a " not in pauta_csv.read_text(encoding="utf-8")

    worksheet = openpyxl.load_workbook(pauta_xlsx, read_only=True)["datos"]
    xlsx_rows = list(worksheet.iter_rows(values_only=True))
    assert list(xlsx_rows[0]) == expected_headers
    xlsx_first = dict(zip(expected_headers, xlsx_rows[1], strict=True))
    assert xlsx_first["Estado"] == "Activo"
    assert xlsx_first["Campaña Principal de Origen"] == (
        "[IND] Campañas Pauta Digital"
    )
    for api_name in (
        "npe03__Contact__r.MailingState",
        "npe03__Contact__r.OtherState",
        "Campa_a_de_origen__c",
    ):
        assert api_name not in expected_headers

    metadata_path = next(Path(path) for path in pauta.artifacts if path.endswith(".json"))
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    assert metadata["variant_id"] == "ind_campanas_pauta_digital"
    assert metadata["soql"] == pauta.soql
    assert metadata["api_name_to_label"]["npe03__Amount__c"] == "Importe"
    assert metadata["api_name_to_label"]["__derived__.age"] == "Edad"
    assert metadata["value_labels"]["npsp__Status__c"]["Active"] == "Activo"
    assert metadata["output_order"][:4] == [
        "npe03__Contact__r.FirstName",
        "npe03__Contact__r.LastName",
        "npe03__Contact__r.Name",
        "npe03__Contact__r.Birthdate",
    ]

    with sqlite3.connect(settings.worker_db_path) as connection:
        rows = connection.execute(
            "SELECT variant_id, soql, artifacts_json FROM report_run_variants ORDER BY id"
        ).fetchall()
        artifact_count = connection.execute(
            "SELECT COUNT(*) FROM report_artifacts"
        ).fetchone()
    assert [row[0] for row in rows] == [
        "ind_campanas_pauta_digital",
        "ind_redes_sociales",
    ]
    assert all(json.loads(row[2]) for row in rows)
    assert artifact_count is not None and artifact_count[0] >= 6


def test_task_23_non_filterable_main_campaign_needs_semantic_clarification(
    tmp_path: Path,
    micaela_task: ExternalTask,
) -> None:
    task_23 = micaela_task.model_copy(update={"id": 23})
    root = Path(__file__).parents[1]
    salesforce = NonFilterableMainCampaignSalesforceClient()

    result, _ = _run(
        tmp_path,
        task_23,
        salesforce,
        mapping_path=root / "config" / "field_mapping.json",
        business_semantics_path=root / "config" / "business_semantics.yaml",
    )

    assert result.status == "needs_clarification"
    assert salesforce.queried_soql == []
    assert (
        "La dimensión Campaña Principal de Origen existe en el reporte Salesforce, "
        "pero no está disponible/filtrable por SOQL con el usuario actual."
        in result.response_text
    )


def test_dry_run_never_queries_salesforce(
    tmp_path: Path,
    micaela_task: ExternalTask,
    fake_salesforce: FakeSalesforceClient,
) -> None:
    result, _ = _run(
        tmp_path,
        micaela_task,
        fake_salesforce,
        mapping_path=write_field_mapping(tmp_path / "mapping.json"),
        dry_run=True,
    )

    assert result.status == "dry_run_completed"
    assert fake_salesforce.queried_soql == []
    assert "Salesforce no fue consultado" in result.response_text


def test_missing_person_relationship_finishes_needs_clarification_and_persists_context(
    tmp_path: Path,
    micaela_task: ExternalTask,
    fake_salesforce: FakeSalesforceClient,
) -> None:
    result, settings = _run(
        tmp_path,
        micaela_task,
        fake_salesforce,
        mapping_path=write_field_mapping(tmp_path / "mapping.json", include_relationship=False),
    )

    assert result.status == "needs_clarification"
    assert result.errors == []
    assert fake_salesforce.queried_soql == []
    assert "¿Qué relación de Opportunity conecta la donación con la persona" in result.response_text
    with sqlite3.connect(settings.worker_db_path) as connection:
        connection.row_factory = sqlite3.Row
        row = connection.execute("SELECT * FROM report_runs ORDER BY id DESC LIMIT 1").fetchone()
    assert row is not None
    assert row["status"] == "needs_clarification"
    assert json.loads(row["request_json"])["person_fields"]
    assert json.loads(row["plan_json"])["needs_clarification"] is True
    assert json.loads(row["warnings_json"])
    assert "Preguntas para Iván" in row["response_text"]


def test_missing_origin_mapping_finishes_needs_clarification(
    tmp_path: Path,
    micaela_task: ExternalTask,
    fake_salesforce: FakeSalesforceClient,
) -> None:
    result, _ = _run(
        tmp_path,
        micaela_task,
        fake_salesforce,
        mapping_path=write_field_mapping(tmp_path / "mapping.json", include_origin=False),
    )

    assert result.status == "needs_clarification"
    assert fake_salesforce.queried_soql == []
    assert "¿Qué campo de Salesforce representa campaña de origen/fuente?" in result.response_text


def test_salesforce_report_creation_is_disabled_by_default_and_failure_is_non_fatal(
    tmp_path: Path,
    micaela_task: ExternalTask,
) -> None:
    mapping_path = write_field_mapping(tmp_path / "mapping.json")
    disabled_client = ReportCreationSalesforceClient()

    disabled_result, _ = _run(
        tmp_path / "disabled",
        micaela_task,
        disabled_client,
        mapping_path=mapping_path,
    )

    assert disabled_result.status == "done_pending_approval"
    assert disabled_client.create_attempts == 0

    enabled_mapping = write_field_mapping(tmp_path / "enabled-mapping.json")
    enabled_client = ReportCreationSalesforceClient()
    enabled_result, _ = _run(
        tmp_path / "enabled",
        micaela_task,
        enabled_client,
        mapping_path=enabled_mapping,
        allow_salesforce_report_create=True,
    )

    assert enabled_result.status == "done_pending_approval"
    assert enabled_client.create_attempts == 1
    assert any("Los archivos locales se conservaron" in value for value in enabled_result.warnings)
    assert any(path.endswith(".csv") for path in enabled_result.artifacts)
