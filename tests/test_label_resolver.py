import json
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import pandas as pd

from sf_report_agent.reports.exporters import export_report
from sf_report_agent.reports.transforms import prepare_export_dataframe
from sf_report_agent.salesforce.label_resolver import SalesforceLabelResolver


def _snapshot() -> dict[str, object]:
    return {
        "objects": {
            "Donation__c": {
                "fields": [
                    {"name": "Amount__c", "label": "Importe", "type": "currency"},
                    {"name": "First_Name__c", "label": "Nombre", "type": "string"},
                    {"name": "Other_Name__c", "label": "Nombre", "type": "string"},
                    {
                        "name": "Contact__c",
                        "label": "Contacto",
                        "type": "reference",
                        "referenceTo": ["Contact"],
                        "relationshipName": "Contact__r",
                    },
                    {
                        "name": "Campa_a_de_origen__c",
                        "label": "Campaña de origen",
                        "type": "reference",
                        "referenceTo": ["Campaign"],
                        "relationshipName": "Campa_a_de_origen__r",
                    },
                    {
                        "name": "Status__c",
                        "label": "Estado",
                        "type": "picklist",
                        "picklistValues": [
                            {"value": "Active", "label": "Activo", "active": True}
                        ],
                    },
                ]
            },
            "Contact": {
                "fields": [
                    {"name": "Name", "label": "Nombre", "type": "string"},
                    {"name": "FirstName", "label": "Nombre", "type": "string"},
                    {
                        "name": "Birthdate",
                        "label": "Fecha de nacimiento",
                        "type": "date",
                    },
                ]
            },
            "Campaign": {
                "fields": [
                    {"name": "Name", "label": "Nombre", "type": "string"},
                ]
            },
        }
    }


def test_resolves_direct_and_relationship_labels_and_disambiguates_duplicates() -> None:
    resolver = SalesforceLabelResolver(_snapshot())

    labels = resolver.resolve(
        "Donation__c",
        ["Amount__c", "Contact__r.Birthdate", "First_Name__c", "Other_Name__c"],
    )

    assert labels["Amount__c"] == "Importe"
    assert labels["Contact__r.Birthdate"] == "Contacto: Fecha de nacimiento"
    assert labels["First_Name__c"] != labels["Other_Name__c"]
    assert labels["First_Name__c"].startswith("Nombre (")


def test_relationship_name_uses_lookup_label_and_picklist_values_use_metadata() -> None:
    resolver = SalesforceLabelResolver(_snapshot())

    labels = resolver.resolve(
        "Donation__c",
        ["Campa_a_de_origen__r.Name", "Contact__r.FirstName"],
    )
    values = resolver.resolve_value_labels("Donation__c", ["Status__c"])

    assert labels["Campa_a_de_origen__r.Name"] == "Campaña de origen"
    assert labels["Contact__r.FirstName"] == "Contacto: Nombre"
    assert values == {"Status__c": {"Active": "Activo"}}


def test_csv_xlsx_use_readable_headers_and_metadata_keeps_api_mapping(tmp_path: Path) -> None:
    dataframe = pd.DataFrame(
        [{"Amount__c": 100, "Contact__r.Birthdate": "1990-01-01"}]
    )
    labels = SalesforceLabelResolver(_snapshot()).resolve(
        "Donation__c", list(dataframe.columns)
    )
    metadata = {"api_name_to_label": labels}

    paths = export_report(
        dataframe.rename(columns=labels),
        task_id=1,
        title="Reporte con labels",
        artifacts_dir=tmp_path,
        output_formats=["csv", "xlsx"],
        metadata=metadata,
        warnings=[],
        generated_at=datetime(2026, 1, 1, tzinfo=UTC),
    )

    csv_path = next(path for path in paths if path.suffix == ".csv")
    xlsx_path = next(path for path in paths if path.suffix == ".xlsx")
    assert csv_path.read_text(encoding="utf-8").splitlines()[0] == (
        "Importe,Contacto: Fecha de nacimiento"
    )
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True)
    assert [cell.value for cell in next(workbook["datos"].iter_rows())] == [
        "Importe",
        "Contacto: Fecha de nacimiento",
    ]
    metadata_values = [row[1].value for row in workbook["metadata"].iter_rows(min_row=2)]
    assert json.dumps(labels, ensure_ascii=False) in metadata_values


def test_csv_and_xlsx_strip_html_anchors(tmp_path: Path) -> None:
    dataframe = pd.DataFrame(
        {"Campaña Principal de Origen": ['<a href="/campaign/1">[IND] Redes Sociales</a>']}
    )

    paths = export_report(
        dataframe,
        task_id=2,
        title="Reporte sin HTML",
        artifacts_dir=tmp_path,
        output_formats=["csv", "xlsx"],
        metadata={},
        warnings=[],
        generated_at=datetime(2026, 1, 1, tzinfo=UTC),
    )

    csv_path = next(path for path in paths if path.suffix == ".csv")
    xlsx_path = next(path for path in paths if path.suffix == ".xlsx")
    assert csv_path.read_text(encoding="utf-8").splitlines()[1] == (
        "[IND] Redes Sociales"
    )
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True)
    assert workbook["datos"]["A2"].value == "[IND] Redes Sociales"


def test_export_preparation_keeps_nullable_age_as_integer(tmp_path: Path) -> None:
    dataframe = pd.DataFrame({"__derived__.age": [36.0, None]})

    result = prepare_export_dataframe(
        dataframe,
        value_labels={},
        output_order=["__derived__.age"],
        integer_fields=["__derived__.age"],
    )

    assert str(result["__derived__.age"].dtype) == "Int64"
    assert result.loc[0, "__derived__.age"] == 36
    assert pd.isna(result.loc[1, "__derived__.age"])

    paths = export_report(
        result.rename(columns={"__derived__.age": "Edad"}),
        task_id=3,
        title="Edades enteras",
        artifacts_dir=tmp_path,
        output_formats=["csv"],
        metadata={},
        warnings=[],
        generated_at=datetime(2026, 1, 1, tzinfo=UTC),
    )

    assert paths[0].read_text(encoding="utf-8").splitlines() == ["Edad", "36", '""']
